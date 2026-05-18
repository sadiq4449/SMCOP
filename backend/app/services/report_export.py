"""PDF / Excel exports for quarterly reports."""

from __future__ import annotations

import json
import logging
from io import BytesIO

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook

from app.models.report import Report

_logger = logging.getLogger(__name__)


def _snapshot_as_dict(raw: object | None) -> dict:
    """Coerce JSON column values that may arrive as dict or serialized string."""
    if raw is None:
        return {}
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str) and raw.strip():
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            return {}
        return parsed if isinstance(parsed, dict) else {}
    return {}


def _pdf_safe(text: str) -> str:
    """FPDF core fonts are latin-1; avoid encoding errors on Urdu or smart quotes."""
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _append_kpi_chart_pdf(pdf: FPDF, snap: object) -> None:
    """Horizontal bar chart of KPI scores (included in exported PDF).

    Uses printable width (``epw``) so bars stay inside margins — fixed coordinates
    previously exceeded page width and caused FPDF to abort (broken PDF for Gov/IE/etc.).
    """
    if not isinstance(snap, dict):
        return
    scores = snap.get("kpi_scores")
    if not isinstance(scores, list) or not scores:
        return
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 13)
    pdf.multi_cell(
        0,
        8,
        _pdf_safe("KPI performance chart"),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.ln(1)
    pdf.set_font("Helvetica", size=9)
    pdf.multi_cell(
        0,
        5,
        _pdf_safe("Bars show score achieved versus maximum for each indicator (from the report snapshot)."),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.ln(3)

    epw = float(getattr(pdf, "epw", pdf.w - pdf.l_margin - pdf.r_margin))
    label_w = min(72.0, epw * 0.40)
    frac_w = 20.0
    gap = 3.0
    slack = 4.0  # reserve space so bar + label + fraction columns stay inside printable width
    bar_total_w = max(36.0, epw - label_w - frac_w - gap - slack)
    bar_x = pdf.l_margin + label_w + gap

    pdf.set_draw_color(210, 213, 219)
    pdf.set_fill_color(37, 99, 235)

    for row in scores:
        if not isinstance(row, dict):
            continue
        if pdf.get_y() > pdf.h - pdf.b_margin - 14:
            pdf.add_page()

        name_raw = row.get("kpi_name")
        name = _pdf_safe(str(name_raw or "Indicator"))[:40]
        score_raw, mx_raw = row.get("score"), row.get("max_score")
        if mx_raw is None:
            mx_raw = 5
        try:
            sc = float(score_raw) if score_raw is not None else 0.0
            mx = float(mx_raw) if mx_raw else 5.0
        except (TypeError, ValueError):
            sc, mx = 0.0, 5.0
        pct = min(1.0, max(0.0, sc / mx)) if mx > 0 else 0.0

        y = pdf.get_y()
        pdf.set_xy(pdf.l_margin, y)
        pdf.set_font("Helvetica", size=9)
        pdf.cell(label_w, 6, name[:32], border=0)
        pdf.rect(bar_x, y, bar_total_w, 5.5, style="D")
        inner_w = bar_total_w * pct
        if inner_w > 0.15:
            pdf.rect(bar_x, y, inner_w, 5.5, style="F")
        pdf.set_xy(bar_x + bar_total_w + 2, y)
        pdf.cell(frac_w, 6, _pdf_safe(f"{sc:g}/{mx:g}")[:14], ln=1)


def report_to_xlsx(report: Report) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Summary"
    ws.append(["Quarter", report.quarter])
    ws.append(["School ID", str(report.school_id)])
    ws.append(["Status", report.status.value])
    ws.append(["Summary", report.summary or ""])
    ws.append(["Recommendations", report.recommendations or ""])
    ws.append(["Principal infrastructure notes", report.principal_infrastructure_notes or ""])
    ws.append(["Principal daily activity notes", report.principal_daily_activity_notes or ""])
    ws.append([])

    snap = _snapshot_as_dict(report.generated_snapshot)
    ws.append(["Generated snapshot (JSON)"])
    ws.append([json.dumps(snap, indent=2)])

    ws2 = wb.create_sheet("KPI rows")
    ws2.append(["kpi_name", "score", "max_score", "pct_of_max"])
    scores = snap.get("kpi_scores")
    if isinstance(scores, list):
        for row in scores:
            if isinstance(row, dict):
                sc_raw, mx_raw = row.get("score"), row.get("max_score")
                if mx_raw is None:
                    mx_raw = 5
                try:
                    s_f = float(sc_raw) if sc_raw is not None else 0.0
                    m_f = float(mx_raw) if mx_raw else 5.0
                except (TypeError, ValueError):
                    s_f, m_f = 0.0, 5.0
                pct = round(100.0 * s_f / m_f, 1) if m_f > 0 else 0.0
                ws2.append([row.get("kpi_name"), row.get("score"), row.get("max_score"), pct])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


class _ReportPDF(FPDF):
    pass


def report_to_pdf(report: Report) -> bytes:
    pdf = _ReportPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(
        0,
        8,
        _pdf_safe(f"Quarterly report — {report.quarter}"),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.ln(4)
    pdf.set_font("Helvetica", size=11)
    pdf.multi_cell(
        0,
        6,
        _pdf_safe(f"School ID: {report.school_id}"),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.multi_cell(
        0,
        6,
        _pdf_safe(f"Status: {report.status.value}"),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Summary")
    pdf.ln()
    pdf.set_font("Helvetica", size=11)
    pdf.multi_cell(
        0,
        6,
        _pdf_safe(report.summary or "(none)"),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Recommendations")
    pdf.ln()
    pdf.set_font("Helvetica", size=11)
    pdf.multi_cell(
        0,
        6,
        _pdf_safe(report.recommendations or "(none)"),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.ln(2)

    snap = _snapshot_as_dict(report.generated_snapshot)
    if snap.get("visit_found"):
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Auto-generated metrics")
        pdf.ln()
        pdf.set_font("Helvetica", size=11)
        pdf.multi_cell(
            0,
            6,
            _pdf_safe(f"Aggregate score: {snap.get('aggregate_score')}"),
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )
        pdf.multi_cell(
            0,
            6,
            _pdf_safe(f"Classroom observations: {snap.get('classroom_observation_count')}"),
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )
        att = snap.get("attendance") if isinstance(snap.get("attendance"), dict) else {}
        pdf.multi_cell(
            0,
            6,
            _pdf_safe(
                f"Attendance window: {att.get('period_start')} → {att.get('period_end')} "
                f"(approved teacher rows: {att.get('approved_teacher_attendance_rows')}, "
                f"student daily rows: {att.get('student_daily_entries')})"
            ),
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )

    try:
        _append_kpi_chart_pdf(pdf, snap)
    except Exception as exc:
        _logger.warning("Skipping KPI chart page in PDF export: %s", exc)

    out = pdf.output(dest="S")
    if isinstance(out, str):
        return out.encode("latin-1")
    return bytes(out)
